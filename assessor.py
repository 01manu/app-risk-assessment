import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import RadarChart, Reference
from datetime import date

# ── LOAD DATA FILES ──────────────────────────────────────────────

def load_json(filepath):
    """Reads a JSON file and returns it as a Python dictionary."""
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


# ── RUN THE ASSESSMENT ───────────────────────────────────────────

def run_assessment(questions_data):
    """
    Walks the assessor through each question interactively.
    Collects answers and calculates scores.
    Returns a list of results.
    """
    print("\n" + "="*60)
    print("  APPLICATION SECURITY RISK ASSESSMENT")
    print("  Based on ISO 27001 | NIST CSF | DORA")
    print("="*60)

    app_name = input("\nEnter application name: ").strip()
    app_owner = input("Enter application owner / team: ").strip()
    data_class = input("Data classification (Public / Internal / Confidential / Secret): ").strip()

    print("\nAnswer each question by typing A, B, C, or D.\n")

    results = []

    for category in questions_data["categories"]:
        print(f"\n{'─'*50}")
        print(f"CATEGORY: {category['name']} (Weight: {category['weight']}%)")
        print(f"{'─'*50}")

        category_scores = []

        for q in category["questions"]:
            print(f"\n{q['id']}: {q['text']}")

            for key, option in q["options"].items():
                print(f"  {key}) {option['text']}")

            # Keep asking until a valid answer is given
            while True:
                answer = input("Your answer (A/B/C/D): ").strip().upper()
                if answer in q["options"]:
                    break
                print("Please enter A, B, C, or D")

            score = q["options"][answer]["score"]
            answer_text = q["options"][answer]["text"]
            category_scores.append(score)

            results.append({
                "question_id": q["id"],
                "category": category["name"],
                "question": q["text"],
                "answer": answer,
                "answer_text": answer_text,
                "score": score,
                "weight": category["weight"]
            })

            # Immediate feedback
            if score == 0:
                print(f"  Good — no risk identified for this control")
            elif score <= 15:
                print(f"  Low risk — minor improvement recommended")
            elif score <= 25:
                print(f"  Medium risk — remediation required")
            else:
                print(f"  HIGH RISK — immediate action required")

    return app_name, app_owner, data_class, results


# ── CALCULATE FINAL SCORE ────────────────────────────────────────

def calculate_score(results, questions_data):
    """
    Calculates the weighted risk score.

    Why weighted? Not all categories are equally important.
    Access Control (25%) matters more than Compliance (15%).
    A weighted score reflects real-world risk priorities.
    """
    category_scores = {}

    # Group scores by category
    for r in results:
        cat = r["category"]
        if cat not in category_scores:
            category_scores[cat] = []
        category_scores[cat].append(r["score"])

    # Calculate average score per category
    category_averages = {}
    for cat, scores in category_scores.items():
        category_averages[cat] = sum(scores) / len(scores)

    # Apply weights to get final score
    # Weight is stored as percentage (25) so divide by 100
    weighted_total = 0
    weights = {c["name"]: c["weight"] for c in questions_data["categories"]}

    for cat, avg in category_averages.items():
        weight = weights.get(cat, 0) / 100
        weighted_total += avg * weight

    # Determine risk level
    if weighted_total >= 30:
        risk_level = "Critical"
        risk_colour = "C00000"
    elif weighted_total >= 20:
        risk_level = "High"
        risk_colour = "FF6600"
    elif weighted_total >= 10:
        risk_level = "Medium"
        risk_colour = "FFB300"
    else:
        risk_level = "Low"
        risk_colour = "375623"

    return round(weighted_total, 1), risk_level, risk_colour, category_averages


# ── BUILD EXCEL REPORT ───────────────────────────────────────────

def build_report(app_name, app_owner, data_class,
                 results, score, risk_level, risk_colour,
                 category_averages, recommendations_data):
    """Generates the Excel assessment report."""

    wb = openpyxl.Workbook()

    RED   = PatternFill("solid", fgColor="FFB3B3")
    AMBER = PatternFill("solid", fgColor="FFE0A0")
    GREEN = PatternFill("solid", fgColor="B3FFB3")
    GREY  = PatternFill("solid", fgColor="F2F2F2")

    def border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    center = Alignment(horizontal="center", vertical="center")
    wrap   = Alignment(wrap_text=True, vertical="top")

    today = date.today()

    # ── SHEET 1: EXECUTIVE SUMMARY ────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    # Title
    ws["B2"] = "Application Security Risk Assessment Report"
    ws["B2"].font = Font(bold=True, size=16, color="1F4E79")
    ws["B2"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws.merge_cells("B2:H2")
    ws["B2"].alignment = center

    # Application details
    details = [
        ("Application Name", app_name),
        ("Application Owner", app_owner),
        ("Data Classification", data_class),
        ("Assessment Date", today.strftime("%d %B %Y")),
        ("Framework", "ISO 27001 | NIST CSF | DORA-aligned"),
    ]

    for i, (label, value) in enumerate(details, start=4):
        ws.cell(row=i, column=2).value = label
        ws.cell(row=i, column=2).font = Font(bold=True)
        ws.cell(row=i, column=2).fill = GREY
        ws.cell(row=i, column=2).border = border()
        ws.cell(row=i, column=3).value = value
        ws.cell(row=i, column=3).border = border()
        ws.merge_cells(f"C{i}:H{i}")

    # Overall risk score
    ws["B10"] = "OVERALL RISK SCORE"
    ws["B10"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["B10"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["B10"].alignment = center
    ws.merge_cells("B10:C10")

    ws["D10"] = f"{score} / 40"
    ws["D10"].font = Font(bold=True, size=18,
                          color=risk_colour)
    ws["D10"].alignment = center

    ws["E10"] = risk_level.upper()
    ws["E10"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["E10"].fill = PatternFill("solid", fgColor=risk_colour)
    ws["E10"].alignment = center
    ws.merge_cells("E10:H10")

    # Category scores table
    ws["B12"] = "Risk Score by Category"
    ws["B12"].font = Font(bold=True, size=12, color="1F4E79")

    headers = ["Category", "Weight", "Score", "Risk Level"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=13, column=i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = center
        c.border = border()

    weights_map = {"Access Control": 25, "Data Protection": 25,
                   "Vulnerability Management": 20,
                   "Incident Response": 15,
                   "Compliance & Certifications": 15}

    for row_i, (cat, avg) in enumerate(category_averages.items(), start=14):
        if avg >= 30:
            cat_level = "Critical"
            cat_fill = RED
        elif avg >= 20:
            cat_level = "High"
            cat_fill = PatternFill("solid", fgColor="FFD0A0")
        elif avg >= 10:
            cat_level = "Medium"
            cat_fill = AMBER
        else:
            cat_level = "Low"
            cat_fill = GREEN

        ws.cell(row=row_i, column=2).value = cat
        ws.cell(row=row_i, column=2).border = border()
        ws.cell(row=row_i, column=3).value = f"{weights_map.get(cat, 0)}%"
        ws.cell(row=row_i, column=3).alignment = center
        ws.cell(row=row_i, column=3).border = border()
        ws.cell(row=row_i, column=4).value = round(avg, 1)
        ws.cell(row=row_i, column=4).alignment = center
        ws.cell(row=row_i, column=4).border = border()
        ws.cell(row=row_i, column=5).value = cat_level
        ws.cell(row=row_i, column=5).fill = cat_fill
        ws.cell(row=row_i, column=5).alignment = center
        ws.cell(row=row_i, column=5).font = Font(bold=True)
        ws.cell(row=row_i, column=5).border = border()

    # Column widths
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    for col in ["F","G","H"]:
        ws.column_dimensions[col].width = 14

    # ── SHEET 2: DETAILED FINDINGS ────────────────────────────────
    ws2 = wb.create_sheet("Detailed Findings")

    headers2 = ["Question ID", "Category", "Question",
                 "Answer", "Risk Score", "Recommendation"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = center
        c.border = border()

    for row_i, r in enumerate(results, start=2):
        # Get recommendation if answer is not A (A = best practice)
        rec = ""
        if r["answer"] != "A":
            qid = r["question_id"]
            ans = r["answer"]
            rec = recommendations_data.get(qid, {}).get(ans, "Review and improve this control")

        # Score colour
        if r["score"] >= 30:
            score_fill = RED
        elif r["score"] >= 20:
            score_fill = PatternFill("solid", fgColor="FFD0A0")
        elif r["score"] >= 10:
            score_fill = AMBER
        else:
            score_fill = GREEN

        ws2.cell(row=row_i, column=1).value = r["question_id"]
        ws2.cell(row=row_i, column=1).alignment = center
        ws2.cell(row=row_i, column=2).value = r["category"]
        ws2.cell(row=row_i, column=3).value = r["question"]
        ws2.cell(row=row_i, column=3).alignment = wrap
        ws2.cell(row=row_i, column=4).value = r["answer_text"]
        ws2.cell(row=row_i, column=4).alignment = wrap
        ws2.cell(row=row_i, column=5).value = r["score"]
        ws2.cell(row=row_i, column=5).fill = score_fill
        ws2.cell(row=row_i, column=5).alignment = center
        ws2.cell(row=row_i, column=5).font = Font(bold=True)
        ws2.cell(row=row_i, column=6).value = rec
        ws2.cell(row=row_i, column=6).alignment = wrap

        for col in range(1, 7):
            ws2.cell(row=row_i, column=col).border = border()

        ws2.row_dimensions[row_i].height = 45

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 35
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 45
    ws2.freeze_panes = "A2"

    # ── SAVE ──────────────────────────────────────────────────────
    safe_name = app_name.replace(" ", "_").replace("/", "-")
    filename = f"risk_assessment_{safe_name}_{today.strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename


# ── MAIN ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    questions_data      = load_json("questions.json")
    recommendations_data = load_json("recommendations.json")

    app_name, app_owner, data_class, results = run_assessment(questions_data)

    score, risk_level, risk_colour, category_averages = calculate_score(
        results, questions_data
    )

    print(f"\n{'='*50}")
    print(f"ASSESSMENT COMPLETE")
    print(f"Application:  {app_name}")
    print(f"Risk Score:   {score} / 40")
    print(f"Risk Level:   {risk_level.upper()}")
    print(f"{'='*50}\n")

    filename = build_report(
        app_name, app_owner, data_class,
        results, score, risk_level, risk_colour,
        category_averages, recommendations_data
    )

    print(f"Report saved: {filename}")
    print("Open the Excel file to see the full assessment report.")